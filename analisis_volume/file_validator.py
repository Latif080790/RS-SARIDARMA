"""
File Validator for Auto Volume Calculator
Validates DXF and Excel/RAB files before processing
"""

import os
from pathlib import Path
from typing import Optional, Dict, Any, List, Tuple
import openpyxl
from openpyxl import load_workbook


class ValidationError(Exception):
    """Custom exception for validation errors"""
    def __init__(self, message: str, suggestions: List[str] = None):
        super().__init__(message)
        self.suggestions = suggestions or []


class FileValidator:
    """
    Validates input files for Auto Volume Calculator:
    - File existence and accessibility
    - File format validation
    - File size limits
    - Encoding validation
    - Required content validation
    """
    
    # File size limits
    MAX_DXF_SIZE_MB = 500  # 500MB max for DXF files
    MAX_EXCEL_SIZE_MB = 50  # 50MB max for Excel files
    MIN_FILE_SIZE_BYTES = 100  # Minimum 100 bytes
    
    # Supported encodings
    SUPPORTED_ENCODINGS = ['utf-8', 'utf-16', 'cp1252', 'latin-1']
    
    def __init__(self, logger=None):
        """
        Initialize file validator
        
        Args:
            logger: Optional ProductionLogger instance for logging
        """
        self.logger = logger
    
    def _log(self, level: str, message: str, **kwargs):
        """Internal logging helper"""
        if self.logger:
            log_func = getattr(self.logger, level, None)
            if log_func:
                log_func(message, **kwargs)
    
    def validate_file_exists(self, file_path: str) -> bool:
        """
        Check if file exists and is accessible
        
        Args:
            file_path: Path to file
            
        Returns:
            True if file exists and readable
            
        Raises:
            ValidationError: If file doesn't exist or not readable
        """
        path = Path(file_path)
        
        if not path.exists():
            suggestions = [
                "Check if the file path is correct",
                "Verify the file hasn't been moved or deleted",
                f"Expected path: {path.absolute()}"
            ]
            self._log("error", f"File not found: {file_path}", category="file_validation")
            raise ValidationError(f"File not found: {file_path}", suggestions)
        
        if not path.is_file():
            suggestions = [f"{file_path} is a directory, not a file"]
            self._log("error", f"Not a file: {file_path}", category="file_validation")
            raise ValidationError(f"Not a file: {file_path}", suggestions)
        
        if not os.access(path, os.R_OK):
            suggestions = [
                "Check file permissions",
                "Close the file if it's open in another program",
                "Run the application with appropriate permissions"
            ]
            self._log("error", f"File not readable: {file_path}", category="file_validation")
            raise ValidationError(f"File not readable: {file_path}", suggestions)
        
        return True
    
    def validate_file_size(
        self,
        file_path: str,
        max_size_mb: Optional[float] = None
    ) -> Tuple[bool, int]:
        """
        Validate file size within acceptable limits
        
        Args:
            file_path: Path to file
            max_size_mb: Maximum size in MB (uses defaults if not specified)
            
        Returns:
            Tuple of (is_valid, file_size_bytes)
            
        Raises:
            ValidationError: If file size is invalid
        """
        path = Path(file_path)
        file_size = path.stat().st_size
        
        # Check minimum size
        if file_size < self.MIN_FILE_SIZE_BYTES:
            suggestions = [
                "File appears to be empty or corrupted",
                "Try re-exporting from source application",
                "Verify file downloaded completely"
            ]
            self._log("error", f"File too small: {file_path} ({file_size} bytes)", category="file_validation")
            raise ValidationError(
                f"File too small: {file_size} bytes (minimum {self.MIN_FILE_SIZE_BYTES} bytes)",
                suggestions
            )
        
        # Check maximum size based on file type
        if max_size_mb is None:
            ext = path.suffix.lower()
            max_size_mb = self.MAX_DXF_SIZE_MB if ext in ['.dxf', '.dwg'] else self.MAX_EXCEL_SIZE_MB
        
        max_size_bytes = max_size_mb * 1024 * 1024
        
        if file_size > max_size_bytes:
            file_size_mb = file_size / (1024 * 1024)
            suggestions = [
                f"File size: {file_size_mb:.1f}MB exceeds limit of {max_size_mb}MB",
                "Try splitting the drawing into smaller files",
                "Remove unnecessary layers or objects",
                "Purge unused blocks and styles"
            ]
            self._log("error", f"File too large: {file_path} ({file_size_mb:.1f}MB)", category="file_validation")
            raise ValidationError(
                f"File too large: {file_size_mb:.1f}MB (max {max_size_mb}MB)",
                suggestions
            )
        
        return True, file_size
    
    def validate_file_extension(
        self,
        file_path: str,
        allowed_extensions: List[str]
    ) -> bool:
        """
        Validate file has correct extension
        
        Args:
            file_path: Path to file
            allowed_extensions: List of allowed extensions (e.g., ['.dxf', '.dwg'])
            
        Returns:
            True if extension is valid
            
        Raises:
            ValidationError: If extension is not allowed
        """
        path = Path(file_path)
        ext = path.suffix.lower()
        allowed_lower = [e.lower() for e in allowed_extensions]
        
        if ext not in allowed_lower:
            suggestions = [
                f"Expected: {', '.join(allowed_extensions)}",
                f"Got: {ext or '(no extension)'}",
                "Verify file format and rename if necessary"
            ]
            self._log("error", f"Invalid extension: {file_path}", category="file_validation")
            raise ValidationError(
                f"Invalid file extension: {ext}. Expected one of {allowed_extensions}",
                suggestions
            )
        
        return True
    
    def validate_dxf_file(self, file_path: str) -> Dict[str, Any]:
        """
        Comprehensive DXF file validation
        
        Args:
            file_path: Path to DXF file
            
        Returns:
            Dict with validation results and file info
            
        Raises:
            ValidationError: If validation fails
        """
        # Step 1: Basic validation
        self.validate_file_exists(file_path)
        self.validate_file_extension(file_path, ['.dxf'])
        is_valid, file_size = self.validate_file_size(file_path, self.MAX_DXF_SIZE_MB)
        
        # Step 2: Check if it's actually a DXF file (should start with 0\r\nSECTION or similar)
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                first_lines = f.read(1000)
                
                # DXF files should contain these markers
                if '0\n' not in first_lines and 'SECTION' not in first_lines:
                    suggestions = [
                        "File may be corrupted or not a valid DXF",
                        "Try opening in CAD software and re-saving",
                        "Export as DXF R2010 or R2013 format"
                    ]
                    self._log("error", f"Invalid DXF format: {file_path}", category="file_validation")
                    raise ValidationError("File doesn't appear to be a valid DXF file", suggestions)
        
        except UnicodeDecodeError:
            # Try binary mode
            try:
                with open(file_path, 'rb') as f:
                    first_bytes = f.read(100)
                    if b'0\n' not in first_bytes and b'SECTION' not in first_bytes:
                        suggestions = [
                            "File encoding not recognized",
                            "Try converting to DXF R2010 ASCII format"
                        ]
                        raise ValidationError("Unable to read DXF file - encoding issue", suggestions)
            except Exception as e:
                self._log("error", f"DXF read error: {file_path} - {str(e)}", category="file_validation")
                raise ValidationError(f"Error reading DXF file: {str(e)}", [])
        
        self._log("info", f"DXF file validated: {file_path} ({file_size / 1024 / 1024:.1f}MB)")
        
        return {
            'valid': True,
            'file_path': file_path,
            'file_size': file_size,
            'file_size_mb': round(file_size / (1024 * 1024), 2),
            'file_type': 'DXF'
        }
    
    def validate_excel_file(self, file_path: str, required_columns: List[str] = None) -> Dict[str, Any]:
        """
        Comprehensive Excel/RAB file validation
        
        Args:
            file_path: Path to Excel file
            required_columns: List of required column names
            
        Returns:
            Dict with validation results and file info
            
        Raises:
            ValidationError: If validation fails
        """
        # Step 1: Basic validation
        self.validate_file_exists(file_path)
        self.validate_file_extension(file_path, ['.xlsx', '.xls'])
        is_valid, file_size = self.validate_file_size(file_path, self.MAX_EXCEL_SIZE_MB)
        
        # Step 2: Try opening with openpyxl
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            sheet_names = wb.sheetnames
            
            if not sheet_names:
                wb.close()
                suggestions = ["File contains no worksheets"]
                self._log("error", f"Empty Excel file: {file_path}", category="file_validation")
                raise ValidationError("Excel file contains no worksheets", suggestions)
            
            # Get first sheet
            ws = wb[sheet_names[0]]
            
            # Check if sheet has data
            if ws.max_row < 2:  # At least header + 1 data row
                wb.close()
                suggestions = [
                    "File appears to be empty",
                    "Ensure RAB data is in the first sheet",
                    "Check if data starts from row 1"
                ]
                self._log("error", f"No data in Excel file: {file_path}", category="file_validation")
                raise ValidationError("Excel file contains no data", suggestions)
            
            # Validate required columns if specified
            if required_columns:
                # Get header row (assume first row)
                header = [cell.value for cell in ws[1]]
                header = [str(h).strip() if h else "" for h in header]
                
                missing_columns = [col for col in required_columns if col not in header]
                
                if missing_columns:
                    wb.close()
                    suggestions = [
                        f"Missing columns: {', '.join(missing_columns)}",
                        f"Found columns: {', '.join(header[:10])}...",
                        "Ensure column names match exactly (case-sensitive)"
                    ]
                    self._log("error", f"Missing columns in Excel: {file_path}", category="file_validation")
                    raise ValidationError(
                        f"Missing required columns: {', '.join(missing_columns)}",
                        suggestions
                    )
            
            row_count = ws.max_row
            wb.close()
            
            self._log("info", f"Excel file validated: {file_path} ({file_size / 1024 / 1024:.1f}MB, {row_count} rows)")
            
            return {
                'valid': True,
                'file_path': file_path,
                'file_size': file_size,
                'file_size_mb': round(file_size / (1024 * 1024), 2),
                'file_type': 'Excel',
                'sheet_names': sheet_names,
                'row_count': row_count
            }
        
        except openpyxl.utils.exceptions.InvalidFileException:
            suggestions = [
                "File may be corrupted",
                "Try opening in Excel and saving as .xlsx",
                "Verify file is not password-protected"
            ]
            self._log("error", f"Invalid Excel file: {file_path}", category="file_validation")
            raise ValidationError("Invalid Excel file format", suggestions)
        
        except PermissionError:
            suggestions = [
                "Close the file if it's open in Excel",
                "Check file permissions",
                "Try running as administrator"
            ]
            self._log("error", f"Permission denied: {file_path}", category="file_validation")
            raise ValidationError("Permission denied - file may be open", suggestions)
        
        except Exception as e:
            self._log("error", f"Excel validation error: {file_path} - {str(e)}", category="file_validation")
            raise ValidationError(f"Error validating Excel file: {str(e)}", [])
    
    def validate_batch(
        self,
        dxf_file: str,
        rab_file: Optional[str] = None
    ) -> Dict[str, Any]:
        """
        Validate a complete set of input files
        
        Args:
            dxf_file: Path to DXF file
            rab_file: Optional path to RAB Excel file
            
        Returns:
            Dict with validation results for all files
            
        Raises:
            ValidationError: If any validation fails
        """
        results = {
            'valid': True,
            'files': {}
        }
        
        # Validate DXF
        try:
            dxf_result = self.validate_dxf_file(dxf_file)
            results['files']['dxf'] = dxf_result
        except ValidationError as e:
            results['valid'] = False
            results['files']['dxf'] = {'valid': False, 'error': str(e), 'suggestions': e.suggestions}
            raise
        
        # Validate RAB if provided
        if rab_file:
            try:
                rab_result = self.validate_excel_file(
                    rab_file,
                    required_columns=['Kode', 'Uraian', 'Volume', 'Satuan']
                )
                results['files']['rab'] = rab_result
            except ValidationError as e:
                results['valid'] = False
                results['files']['rab'] = {'valid': False, 'error': str(e), 'suggestions': e.suggestions}
                raise
        
        return results


if __name__ == '__main__':
    # Demo usage
    print("Testing File Validator...")
    
    from production_logger import create_logger
    
    logger = create_logger("FileValidatorTest", log_level="DEBUG")
    validator = FileValidator(logger)
    
    # Test with non-existent file
    try:
        validator.validate_dxf_file("non_existent.dxf")
    except ValidationError as e:
        print(f"\n✓ Caught expected error: {e}")
        print(f"  Suggestions: {e.suggestions}")
    
    # Test with this Python file (wrong extension)
    try:
        validator.validate_dxf_file(__file__)
    except ValidationError as e:
        print(f"\n✓ Caught expected error: {e}")
        print(f"  Suggestions: {e.suggestions}")
    
    logger.close()
    print("\n✓ File validator test complete")
