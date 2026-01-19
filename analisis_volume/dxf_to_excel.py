"""
DXF to Excel Converter
Workflow: DXF File → Extract Data → Calculate Volume → Populate Excel Template
"""

import os
import sys
from pathlib import Path

# Add analisis_volume to path
sys.path.insert(0, os.path.dirname(__file__))

from dwg_reader import DXFReader
from auto_volume_calculator import AutoVolumeCalculator
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime


class DXFToExcelConverter:
    """Convert DXF data ke Excel template"""
    
    def __init__(self, dxf_file: str, template_file: str, output_file: str):
        self.dxf_file = dxf_file
        self.template_file = template_file
        self.output_file = output_file
        self.items = []
        
    def extract_from_dxf(self) -> bool:
        """Extract data dari DXF file"""
        print("\n" + "="*70)
        print("STEP 1: EXTRACT DATA FROM DXF")
        print("="*70)
        
        reader = DXFReader(self.dxf_file)
        
        if not reader.load_file():
            return False
        
        # Extract all data
        reader.extract_layers()
        reader.extract_texts()
        reader.extract_dimensions()
        reader.extract_polylines()
        reader.extract_lines()
        reader.extract_circles()
        reader.extract_hatches()
        
        print("\n" + "="*70)
        print("STEP 2: AUTO CALCULATE VOLUMES")
        print("="*70)
        
        # Calculate volumes
        calculator = AutoVolumeCalculator(reader.data)
        self.items = calculator.calculate_all_volumes()
        
        # Show summary
        summary = calculator.get_summary_by_category()
        print("\nVolume Summary:")
        for kategori, stats in summary.items():
            print(f"  • {kategori.upper():15} : {stats['count']:3} items, "
                  f"Total: {stats['total_volume']:10.2f} m³")
        
        return True
    
    def populate_template(self) -> bool:
        """Populate Excel template dengan data"""
        print("\n" + "="*70)
        print("STEP 3: POPULATE EXCEL TEMPLATE")
        print("="*70)
        
        if not os.path.exists(self.template_file):
            print(f"✗ Template file tidak ditemukan: {self.template_file}")
            return False
        
        try:
            # Load template
            wb = load_workbook(self.template_file)
            print(f"✓ Template loaded: {os.path.basename(self.template_file)}")
            
            # Group items by category
            grouped_items = {
                'struktur': [],
                'arsitektur': [],
                'mep': []
            }
            
            kategori_mapping = {
                'kolom': 'struktur',
                'balok': 'struktur',
                'plat': 'struktur',
                'sloof': 'struktur',
                'pondasi': 'struktur',
                'ring': 'struktur',
                'tangga': 'struktur',
                'dinding': 'arsitektur',
                'pintu': 'arsitektur',
                'jendela': 'arsitektur',
                'lantai': 'arsitektur',
                'plafon': 'arsitektur',
                'atap': 'arsitektur',
            }
            
            for item in self.items:
                kategori = item.get('kategori', 'unknown')
                group = kategori_mapping.get(kategori, 'arsitektur')
                grouped_items[group].append(item)
            
            # Populate each sheet
            sheets_map = {
                'struktur': 'STRUKTUR',
                'arsitektur': 'ARSITEKTUR',
                'mep': 'MEP'
            }
            
            fill_green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            
            for group, sheet_name in sheets_map.items():
                if sheet_name not in wb.sheetnames:
                    print(f"  ⚠ Sheet {sheet_name} tidak ditemukan, skip...")
                    continue
                
                ws = wb[sheet_name]
                items = grouped_items[group]
                
                if not items:
                    print(f"  • {sheet_name}: No items to populate")
                    continue
                
                # Find starting row (after header)
                start_row = 6  # Default starting row after header
                
                # Populate items (WITH ENHANCED COLUMNS)
                row = start_row
                for idx, item in enumerate(items, 1):
                    ws.cell(row=row, column=1).value = idx  # No
                    ws.cell(row=row, column=2).value = item.get('kode', '')  # Kode
                    ws.cell(row=row, column=3).value = item.get('item', '')  # Item
                    ws.cell(row=row, column=4).value = item.get('lantai', '')  # Lantai
                    ws.cell(row=row, column=5).value = item.get('grid', '')  # Lokasi/Grid
                    ws.cell(row=row, column=6).value = item.get('panjang', 0)  # Panjang
                    ws.cell(row=row, column=7).value = item.get('lebar', 0)  # Lebar
                    ws.cell(row=row, column=8).value = item.get('tinggi', 0) if item.get('tinggi') else ''  # Tinggi
                    ws.cell(row=row, column=9).value = item.get('jumlah', 1)  # Jumlah
                    ws.cell(row=row, column=10).value = item.get('satuan', 'm3')  # Satuan
                    ws.cell(row=row, column=11).value = item.get('volume', 0)  # Volume
                    ws.cell(row=row, column=12).value = f"Auto: {item.get('method', 'DXF')}"  # Metode
                    
                    # Apply green fill to auto-populated rows
                    for col in range(1, 13):
                        ws.cell(row=row, column=col).fill = fill_green
                    
                    row += 1
                
                print(f"  ✓ {sheet_name}: {len(items)} items populated")
            
            # Save output
            wb.save(self.output_file)
            print(f"\n✓ Excel file saved: {self.output_file}")
            
            return True
            
        except Exception as e:
            print(f"✗ Error populating template: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def run_conversion(self) -> bool:
        """Run full conversion workflow"""
        print("\n" + "╔" + "="*68 + "╗")
        print("║" + " "*18 + "DXF TO EXCEL CONVERTER" + " "*28 + "║")
        print("╚" + "="*68 + "╝")
        print(f"\nTimestamp: {datetime.now().strftime('%d-%m-%Y %H:%M:%S')}")
        
        # Step 1: Extract from DXF
        if not self.extract_from_dxf():
            print("\n✗ Failed to extract from DXF")
            return False
        
        # Step 2: Populate template
        if not self.populate_template():
            print("\n✗ Failed to populate template")
            return False
        
        print("\n" + "="*70)
        print("✓ CONVERSION COMPLETED SUCCESSFULLY")
        print("="*70)
        print(f"\nOutput file: {self.output_file}")
        print("\nNext steps:")
        print("1. Open the Excel file and review auto-populated data")
        print("2. Adjust/correct any values if needed")
        print("3. Add missing items manually if any")
        print("4. Save as 'Volume_dari_Gambar.xlsx'")
        print("5. Run analysis: RUN_ANALISIS.bat")
        print()
        
        return True


def main():
    """Main function"""
    # Default paths
    base_dir = Path(__file__).parent.parent
    
    dxf_file = base_dir / "drawing" / "ars" / "20251108_Plan RS Sari Dharma.dxf"
    template_file = base_dir / "Volume_dari_Gambar_TEMPLATE.xlsx"
    output_file = base_dir / "Volume_dari_Gambar_AUTO.xlsx"
    
    # Check if DXF file exists
    if not dxf_file.exists():
        print(f"\n✗ DXF file tidak ditemukan: {dxf_file}")
        print("\nPastikan file DWG sudah diconvert ke DXF!")
        print("\nCara convert:")
        print("1. Buka DWG di AutoCAD")
        print("2. File > Save As > DXF")
        print("3. Simpan dengan nama yang sama")
        print("\nAtau gunakan ODA File Converter (gratis)")
        return False
    
    # Check if template exists
    if not template_file.exists():
        print(f"\n✗ Template tidak ditemukan: {template_file}")
        print("\nJalankan dulu: GENERATE_TEMPLATE.bat")
        return False
    
    # Run conversion
    converter = DXFToExcelConverter(
        str(dxf_file),
        str(template_file),
        str(output_file)
    )
    
    return converter.run_conversion()


if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)
