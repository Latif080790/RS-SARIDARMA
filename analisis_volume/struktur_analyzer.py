"""
STRUKTUR ANALYZER - Enhanced Analysis for Structural Work
Provides detailed matching and comparison for structural items
"""

import pandas as pd
import re
from typing import Dict, List, Tuple, Optional
from difflib import SequenceMatcher


class StrukturAnalyzer:
    """Enhanced analyzer specifically for structural work items"""
    
    # Kategori pekerjaan struktur
    CATEGORIES = {
        'TANAH': ['galian', 'urugan', 'timbunan', 'pemadatan', 'pembuangan'],
        'PONDASI': ['pondasi', 'tiang pancang', 'foot plate', 'pile cap', 'poer'],
        'BETON': ['beton', 'cor', 'concrete', 'ready mix', 'plat', 'sloop', 'poor', 'lantai kerja'],
        'BEKISTING': ['bekisting', 'formwork', 'cetakan', 'mal'],
        'PEMBESIAN': ['besi', 'baja', 'tulangan', 'reinforcement', 'wiremesh'],
        'BALOK': ['balok', 'beam', 'ring balk'],
        'KOLOM': ['kolom', 'column', 'pilar'],
        'DINDING': ['dinding', 'wall', 'shear wall'],
        'TANGGA': ['tangga', 'stair', 'bordes'],
        'ATAP': ['atap', 'roof', 'kuda-kuda', 'rangka atap'],
    }
    
    # Material specifications yang harus di-extract
    SPEC_PATTERNS = {
        'beton_grade': r'k-?\s*(\d+)',  # K-225, K-300, etc
        'beton_fc': r'fc\s*(\d+)',       # fc 25, fc 30
        'diameter_besi': r'd\s*(\d+)|diameter\s*(\d+)|ø\s*(\d+)',  # D13, diameter 16
        'dimensi': r'(\d+\.?\d*)\s*x\s*(\d+\.?\d*)',  # 40x60, 30x40
        'tebal': r't\s*=?\s*(\d+)|tebal\s*(\d+)',  # t=12, tebal 15
        'panjang': r'p\s*=?\s*(\d+)|panjang\s*(\d+)',  # p=6000
    }
    
    def __init__(self):
        """Initialize struktur analyzer"""
        self.matched_items = []
        self.unmatched_gambar = []
        self.unmatched_rab = []
        self.category_summary = {}
    
    def extract_specifications(self, text: str) -> Dict[str, str]:
        """Extract technical specifications from item description
        
        Args:
            text: Item description text
            
        Returns:
            Dictionary of extracted specifications
        """
        specs = {}
        text_lower = str(text).lower()
        
        for spec_name, pattern in self.SPEC_PATTERNS.items():
            match = re.search(pattern, text_lower)
            if match:
                # Get the first non-None group
                value = next((g for g in match.groups() if g is not None), None)
                if value:
                    specs[spec_name] = value
        
        return specs
    
    def categorize_item(self, item_text: str) -> Optional[str]:
        """Categorize structural item
        
        Args:
            item_text: Item description
            
        Returns:
            Category name or None
        """
        text_lower = str(item_text).lower()
        
        for category, keywords in self.CATEGORIES.items():
            for keyword in keywords:
                if keyword in text_lower:
                    return category
        
        return 'LAIN-LAIN'
    
    def calculate_similarity(self, text1: str, text2: str, 
                           check_specs: bool = True) -> Tuple[float, Dict]:
        """Calculate similarity with specification matching
        
        Args:
            text1: First text
            text2: Second text
            check_specs: Whether to check specifications
            
        Returns:
            Tuple of (similarity_score, match_details)
        """
        text1_clean = str(text1).lower().strip()
        text2_clean = str(text2).lower().strip()
        
        # Exact match
        if text1_clean == text2_clean:
            return 1.0, {'type': 'exact', 'specs_match': True}
        
        # Extract specifications
        specs1 = self.extract_specifications(text1) if check_specs else {}
        specs2 = self.extract_specifications(text2) if check_specs else {}
        
        match_details = {
            'type': 'fuzzy',
            'specs_match': True,
            'specs_compared': [],
            'specs_different': []
        }
        
        # Check specification compatibility
        if check_specs and specs1 and specs2:
            # Compare critical specs
            for spec_name in ['beton_grade', 'beton_fc', 'diameter_besi']:
                if spec_name in specs1 and spec_name in specs2:
                    match_details['specs_compared'].append(spec_name)
                    if specs1[spec_name] != specs2[spec_name]:
                        match_details['specs_match'] = False
                        match_details['specs_different'].append({
                            'spec': spec_name,
                            'value1': specs1[spec_name],
                            'value2': specs2[spec_name]
                        })
        
        # If critical specs don't match, return low similarity
        if not match_details['specs_match']:
            return 0.3, match_details
        
        # Calculate base similarity
        # Check containment first
        if text1_clean in text2_clean or text2_clean in text1_clean:
            base_similarity = 0.85
        else:
            # Use sequence matcher
            base_similarity = SequenceMatcher(None, text1_clean, text2_clean).ratio()
        
        # Boost similarity if specs match
        if match_details['specs_compared'] and match_details['specs_match']:
            base_similarity = min(1.0, base_similarity + 0.1)
        
        match_details['base_similarity'] = base_similarity
        
        return base_similarity, match_details
    
    def match_items(self, gambar_df: pd.DataFrame, rab_df: pd.DataFrame,
                   min_similarity: float = 0.75) -> pd.DataFrame:
        """Match items from gambar with RAB
        
        Args:
            gambar_df: DataFrame from volume gambar (struktur sheet)
            rab_df: DataFrame from RAB struktur
            min_similarity: Minimum similarity threshold
            
        Returns:
            DataFrame with matching results
        """
        print("\n" + "="*80)
        print("DETAILED STRUKTUR ANALYSIS")
        print("="*80)
        
        results = []
        
        # Clean RAB data
        rab_clean = rab_df[pd.notna(rab_df['PEKERJAAN'])].copy()
        rab_clean = rab_clean[pd.notna(rab_clean['VOLUME'])]
        rab_clean = rab_clean[~rab_clean['PEKERJAAN'].astype(str).str.match(r'^[IVX]+$')]  # Remove roman numerals
        rab_clean = rab_clean[~rab_clean['PEKERJAAN'].astype(str).str.contains('Sub Total|TOTAL', case=False, na=False)]
        
        print(f"\n📊 Data Overview:")
        print(f"   Gambar items: {len(gambar_df)}")
        print(f"   RAB items: {len(rab_clean)}")
        print(f"   Similarity threshold: {min_similarity*100}%")
        
        # Categorize items
        print(f"\n📂 Categorizing items...")
        
        for _, gambar_row in gambar_df.iterrows():
            item_gambar = str(gambar_row.get('Item', gambar_row.get('Uraian', '')))
            vol_gambar = float(gambar_row.get('Volume', 0))
            satuan_gambar = str(gambar_row.get('Satuan', ''))
            lokasi_gambar = str(gambar_row.get('Lokasi', gambar_row.get('Lantai', '')))
            
            if not item_gambar or item_gambar == 'nan' or vol_gambar == 0:
                continue
            
            # Categorize
            category = self.categorize_item(item_gambar)
            specs_gambar = self.extract_specifications(item_gambar)
            
            # Find best match in RAB
            best_match = None
            best_similarity = 0.0
            best_details = {}
            
            for _, rab_row in rab_clean.iterrows():
                item_rab = str(rab_row['PEKERJAAN'])
                
                similarity, details = self.calculate_similarity(
                    item_gambar, item_rab, check_specs=True
                )
                
                if similarity > best_similarity and similarity >= min_similarity:
                    best_similarity = similarity
                    best_match = rab_row
                    best_details = details
            
            # Build result
            result = {
                'No': gambar_row.get('No', ''),
                'Kategori': category,
                'Item_Gambar': item_gambar,
                'Lokasi': lokasi_gambar,
                'Volume_Gambar': vol_gambar,
                'Satuan_Gambar': satuan_gambar,
                'Spesifikasi_Gambar': str(specs_gambar) if specs_gambar else '-',
            }
            
            if best_match is not None:
                vol_rab = float(best_match['VOLUME'])
                harga_satuan = float(best_match.get('UNIT PRICE', 0))
                total_rab = float(best_match.get('HARGA', 0))
                
                selisih = vol_gambar - vol_rab
                selisih_persen = (selisih / vol_rab * 100) if vol_rab > 0 else 0
                
                # Determine status
                if abs(selisih_persen) <= 5:
                    status = '✓ OK'
                elif abs(selisih_persen) <= 10:
                    status = '⚠ MINOR'
                elif abs(selisih_persen) <= 25:
                    status = '⚠ WARNING'
                else:
                    status = '❌ MAJOR'
                
                result.update({
                    'Item_RAB': str(best_match['PEKERJAAN']),
                    'Volume_RAB': vol_rab,
                    'Satuan_RAB': str(best_match.get('UNIT', '')),
                    'Selisih_Volume': selisih,
                    'Selisih_%': selisih_persen,
                    'Status': status,
                    'Similarity_%': best_similarity * 100,
                    'Specs_Match': 'YES' if best_details.get('specs_match', True) else 'NO',
                    'Harga_Satuan_RAB': harga_satuan,
                    'Total_RAB': total_rab,
                    'Dampak_Biaya': abs(selisih * harga_satuan) if harga_satuan > 0 else 0,
                    'Match_Type': best_details.get('type', 'fuzzy'),
                })
                
                self.matched_items.append(result)
            else:
                # Unmatched item
                result.update({
                    'Item_RAB': '❌ NOT FOUND IN RAB',
                    'Volume_RAB': 0,
                    'Satuan_RAB': '-',
                    'Selisih_Volume': vol_gambar,
                    'Selisih_%': 0,
                    'Status': '❌ MISSING',
                    'Similarity_%': 0,
                    'Specs_Match': '-',
                    'Harga_Satuan_RAB': 0,
                    'Total_RAB': 0,
                    'Dampak_Biaya': 0,
                    'Match_Type': 'not_found',
                })
                
                self.unmatched_gambar.append(result)
            
            results.append(result)
        
        # Find RAB items not in gambar
        print(f"\n🔍 Checking for RAB items not in gambar...")
        gambar_items_matched = set([r['Item_RAB'] for r in self.matched_items])
        
        for _, rab_row in rab_clean.iterrows():
            item_rab = str(rab_row['PEKERJAAN'])
            if item_rab not in gambar_items_matched:
                self.unmatched_rab.append({
                    'Item': item_rab,
                    'Volume': rab_row['VOLUME'],
                    'Satuan': rab_row.get('UNIT', ''),
                    'Harga_Satuan': rab_row.get('UNIT PRICE', 0),
                    'Total': rab_row.get('HARGA', 0),
                    'Kategori': self.categorize_item(item_rab),
                })
        
        # Create DataFrame
        results_df = pd.DataFrame(results)
        
        # Generate summary
        self._generate_summary(results_df)
        
        return results_df
    
    def _generate_summary(self, results_df: pd.DataFrame):
        """Generate category-wise summary"""
        print("\n" + "="*80)
        print("SUMMARY BY CATEGORY")
        print("="*80)
        
        if not results_df.empty and 'Kategori' in results_df.columns:
            category_groups = results_df.groupby('Kategori')
            
            for category, group in category_groups:
                matched = len(group[group['Status'].str.contains('OK|MINOR|WARNING', na=False)])
                missing = len(group[group['Status'].str.contains('MISSING', na=False)])
                major_diff = len(group[group['Status'].str.contains('MAJOR', na=False)])
                
                total_dampak = group['Dampak_Biaya'].sum() if 'Dampak_Biaya' in group.columns else 0
                
                self.category_summary[category] = {
                    'total_items': len(group),
                    'matched': matched,
                    'missing': missing,
                    'major_diff': major_diff,
                    'total_dampak_biaya': total_dampak,
                }
                
                print(f"\n📂 {category}:")
                print(f"   Total items: {len(group)}")
                print(f"   Matched: {matched}")
                print(f"   Missing in RAB: {missing}")
                print(f"   Major difference: {major_diff}")
                if total_dampak > 0:
                    print(f"   Total cost impact: Rp {total_dampak:,.0f}")
        
        print("\n" + "="*80)
        print(f"\n📊 Overall Statistics:")
        print(f"   Total matched: {len(self.matched_items)}")
        print(f"   Items in Gambar not in RAB: {len(self.unmatched_gambar)}")
        print(f"   Items in RAB not in Gambar: {len(self.unmatched_rab)}")
    
    def generate_detail_report(self, output_file: str):
        """Generate detailed Excel report
        
        Args:
            output_file: Path to output Excel file
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
        ok_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        warning_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        error_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Sheet 1: Summary
        ws_summary = wb.create_sheet("Summary")
        ws_summary.append(["STRUKTUR ANALYSIS SUMMARY"])
        ws_summary.append([])
        ws_summary.append(["Category", "Total Items", "Matched", "Missing", "Major Diff", "Cost Impact (Rp)"])
        
        for row_idx, (category, data) in enumerate(self.category_summary.items(), start=4):
            ws_summary.append([
                category,
                data['total_items'],
                data['matched'],
                data['missing'],
                data['major_diff'],
                data['total_dampak_biaya']
            ])
        
        # Sheet 2: Matched Items
        if self.matched_items:
            ws_matched = wb.create_sheet("Matched Items")
            if self.matched_items:
                df_matched = pd.DataFrame(self.matched_items)
                for r_idx, row in enumerate([df_matched.columns.tolist()] + df_matched.values.tolist(), start=1):
                    ws_matched.append(row)
        
        # Sheet 3: Missing in RAB
        if self.unmatched_gambar:
            ws_missing = wb.create_sheet("Missing in RAB")
            df_missing = pd.DataFrame(self.unmatched_gambar)
            for r_idx, row in enumerate([df_missing.columns.tolist()] + df_missing.values.tolist(), start=1):
                ws_missing.append(row)
        
        # Sheet 4: RAB Not in Gambar
        if self.unmatched_rab:
            ws_rab_only = wb.create_sheet("RAB Not in Gambar")
            df_rab_only = pd.DataFrame(self.unmatched_rab)
            for r_idx, row in enumerate([df_rab_only.columns.tolist()] + df_rab_only.values.tolist(), start=1):
                ws_rab_only.append(row)
        
        # Save
        wb.save(output_file)
        print(f"\n✅ Detailed report saved: {output_file}")


# Helper function for easy use
def analyze_struktur_detail(gambar_file: str, rab_file: str, output_dir: str = "output/reports") -> pd.DataFrame:
    """Convenience function for detailed struktur analysis
    
    Args:
        gambar_file: Path to volume gambar file
        rab_file: Path to RAB struktur file
        output_dir: Output directory for reports
        
    Returns:
        DataFrame with analysis results
    """
    import os
    os.makedirs(output_dir, exist_ok=True)
    
    # Read gambar
    print("Reading volume from gambar...")
    gambar_df = pd.read_excel(gambar_file, sheet_name='STRUKTUR', header=4)
    gambar_df = gambar_df[gambar_df.columns[:10]]
    gambar_df.columns = ['No', 'Item', 'Lokasi', 'Panjang', 'Lebar', 'Tinggi', 'Jumlah', 'Satuan', 'Volume', 'Rumus']
    gambar_df = gambar_df[pd.notna(gambar_df['Item'])]
    gambar_df = gambar_df[pd.notna(gambar_df['Volume'])]
    
    # Read RAB
    print("Reading RAB struktur...")
    rab_df = pd.read_excel(rab_file, sheet_name='BOQ STRUKTUR', header=6)
    
    # Analyze
    analyzer = StrukturAnalyzer()
    results_df = analyzer.match_items(gambar_df, rab_df, min_similarity=0.75)
    
    # Generate report
    timestamp = pd.Timestamp.now().strftime("%Y%m%d_%H%M%S")
    output_file = os.path.join(output_dir, f"STRUKTUR_ANALYSIS_DETAIL_{timestamp}.xlsx")
    analyzer.generate_detail_report(output_file)
    
    return results_df


if __name__ == "__main__":
    # Test the analyzer
    import sys
    import os
    sys.path.insert(0, '.')
    
    gambar_file = "output/volumes/Volume_dari_Gambar_AUTO.xlsx"
    rab_file = "rab/str/BOQ-Dokumen Struktur.xlsx"
    
    if os.path.exists(gambar_file) and os.path.exists(rab_file):
        results = analyze_struktur_detail(gambar_file, rab_file)
        print(f"\n✅ Analysis complete! Results: {len(results)} items")
    else:
        print("Files not found. Please check paths.")
