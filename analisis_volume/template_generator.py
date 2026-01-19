"""
Generator Template Excel untuk Input Volume dari Gambar DED
Template ini akan diisi manual berdasarkan pembacaan gambar
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime


class VolumeTemplateGenerator:
    """Generator template Excel untuk input volume"""
    
    def __init__(self, output_path: str):
        self.output_path = output_path
        self.wb = Workbook()
        
        # Define colors
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.subheader_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
        self.kategori_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        
        # Define fonts
        self.header_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
        self.kategori_font = Font(name='Calibri', size=11, bold=True)
        self.normal_font = Font(name='Calibri', size=10)
        
        # Define border
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
    def create_struktur_sheet(self):
        """Buat sheet untuk pekerjaan struktur"""
        ws = self.wb.create_sheet("STRUKTUR", 0)
        
        # Title
        ws.merge_cells('A1:J1')
        cell = ws['A1']
        cell.value = "VOLUME PEKERJAAN STRUKTUR - DARI GAMBAR DED"
        cell.font = Font(name='Calibri', size=14, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = self.header_fill
        
        # Info
        ws['A2'] = "Project:"
        ws['B2'] = "RS Sari Dharma"
        ws['A3'] = "Tanggal:"
        ws['B3'] = datetime.now().strftime("%d-%m-%Y")
        
        # Header columns - Row 5 (WITH ENHANCED DETAILS)
        headers = ['No', 'Kode', 'Item Pekerjaan', 'Lantai', 'Lokasi/As Grid', 'Panjang (m)', 'Lebar (m)', 'Tinggi (m)', 'Jumlah', 'Satuan', 'Volume', 'Metode']
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col_num)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.thin_border
        
        # Set column widths
        ws.column_dimensions['A'].width = 5   # No
        ws.column_dimensions['B'].width = 8   # Kode
        ws.column_dimensions['C'].width = 30  # Item
        ws.column_dimensions['D'].width = 10  # Lantai
        ws.column_dimensions['E'].width = 15  # Lokasi/Grid
        ws.column_dimensions['F'].width = 10  # Panjang
        ws.column_dimensions['G'].width = 10  # Lebar
        ws.column_dimensions['H'].width = 10  # Tinggi
        ws.column_dimensions['I'].width = 8   # Jumlah
        ws.column_dimensions['J'].width = 8   # Satuan
        ws.column_dimensions['K'].width = 12  # Volume
        ws.column_dimensions['L'].width = 15  # Metode
        
        # Kategori dan contoh data
        row = 6
        
        # ENHANCED CATEGORIES WITH DETAILED BREAKDOWN
        categories = [
            {
                'name': 'A. PEKERJAAN TANAH & PONDASI',
                'items': [
                    ('GT', 'Galian Tanah Pondasi', 'Basement', ''),
                    ('UP', 'Urugan Pasir Bawah Pondasi', 'Basement', ''),
                    ('UT', 'Urugan Tanah Kembali', 'All', ''),
                    ('PT', 'Pemadatan Tanah', 'All', '')
                ]
            },
            {
                'name': 'B. PONDASI & SLOOF',
                'items': [
                    ('P1', 'Pondasi Footplate P1 (100x100x50)', 'Basement', 'Grid A-D'),
                    ('P2', 'Pondasi Footplate P2 (120x120x60)', 'Basement', 'Grid E-H'),
                    ('P3', 'Pondasi Footplate P3 (150x150x70)', 'Basement', 'Grid I-L'),
                    ('S1', 'Sloof S1 K-225 (20/30)', 'Basement', 'As A-L'),
                    ('S2', 'Sloof S2 K-225 (15/25)', 'Basement', 'As 1-10')
                ]
            },
            {
                'name': 'C1. KOLOM BASEMENT',
                'items': [
                    ('K1', 'Kolom K1 K-300 (70x70)', 'Basement', 'As A1-A4'),
                    ('K2', 'Kolom K2 K-300 (65x65)', 'Basement', 'As B1-B6'),
                    ('K3', 'Kolom K3 K-300 (50x50)', 'Basement', 'As C1-C8'),
                    ('K4', 'Kolom K4 Praktis (13/13)', 'Basement', 'Partisi')
                ]
            },
            {
                'name': 'C2. KOLOM LANTAI 1',
                'items': [
                    ('K1', 'Kolom K1 K-300 (70x70)', 'Lantai 1', 'As A1-A4'),
                    ('K2', 'Kolom K2 K-300 (65x65)', 'Lantai 1', 'As B1-B6'),
                    ('K3', 'Kolom K3 K-300 (50x50)', 'Lantai 1', 'As C1-C8'),
                    ('K4', 'Kolom K4 Praktis (13/13)', 'Lantai 1', 'Partisi')
                ]
            },
            {
                'name': 'C3. KOLOM LANTAI 2',
                'items': [
                    ('K1', 'Kolom K1 K-300 (70x70)', 'Lantai 2', 'As A1-A4'),
                    ('K2', 'Kolom K2 K-300 (65x65)', 'Lantai 2', 'As B1-B6'),
                    ('K3', 'Kolom K3 K-300 (50x50)', 'Lantai 2', 'As C1-C8')
                ]
            },
            {
                'name': 'D1. BALOK LANTAI 1',
                'items': [
                    ('B1', 'Balok B1 Induk K-300 (20/35)', 'Lantai 1', 'As A-D'),
                    ('B2', 'Balok B2 Anak K-300 (15/20)', 'Lantai 1', 'As E-H'),
                    ('B3', 'Balok B3 Ring K-225 (13/15)', 'Lantai 1', 'Keliling'),
                    ('B4', 'Balok B4 Latei K-225 (13/13)', 'Lantai 1', 'Pintu/Jendela')
                ]
            },
            {
                'name': 'D2. BALOK LANTAI 2',
                'items': [
                    ('B1', 'Balok B1 Induk K-300 (20/35)', 'Lantai 2', 'As A-D'),
                    ('B2', 'Balok B2 Anak K-300 (15/20)', 'Lantai 2', 'As E-H'),
                    ('B3', 'Balok B3 Ring K-225 (13/15)', 'Lantai 2', 'Keliling')
                ]
            },
            {
                'name': 'E. PLAT LANTAI',
                'items': [
                    ('PL1', 'Plat Lantai 1 K-300 t=12cm', 'Lantai 1', 'Zona A'),
                    ('PL1', 'Plat Lantai 1 K-300 t=12cm', 'Lantai 1', 'Zona B'),
                    ('PL2', 'Plat Lantai 2 K-300 t=12cm', 'Lantai 2', 'Zona A'),
                    ('PL2', 'Plat Lantai 2 K-300 t=12cm', 'Lantai 2', 'Zona B'),
                    ('PL3', 'Plat Atap K-300 t=10cm', 'Atap', 'Full'),
                    ('PT', 'Plat Tangga K-300', 'Tangga', ''),
                    ('PB', 'Plat Bordes K-300', 'Tangga', '')
                ]
            }
        ]
        
        no = 1
        for category in categories:
            # Kategori header
            ws.merge_cells(f'A{row}:L{row}')
            cell = ws[f'A{row}']
            cell.value = category['name']
            cell.font = self.kategori_font
            cell.fill = self.kategori_fill
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = self.thin_border
            row += 1
            
            # Items
            for item_data in category['items']:
                # Handle both tuple and string format
                if isinstance(item_data, tuple):
                    kode, item, lantai, lokasi = item_data
                else:
                    kode, item, lantai, lokasi = '', item_data, '', ''
                
                ws.cell(row=row, column=1).value = no  # No
                ws.cell(row=row, column=2).value = kode  # Kode
                ws.cell(row=row, column=3).value = item  # Item
                ws.cell(row=row, column=4).value = lantai  # Lantai
                ws.cell(row=row, column=5).value = lokasi  # Lokasi/Grid
                
                # Formula untuk volume (P x L x T x Jumlah)
                formula = f"=F{row}*G{row}*H{row}*I{row}"
                ws.cell(row=row, column=11).value = formula
                
                # Apply styles
                for col in range(1, 13):
                    cell = ws.cell(row=row, column=col)
                    cell.border = self.thin_border
                    if col in [1, 2, 4, 9, 10]:  # No, Kode, Lantai, Jumlah, Satuan
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    else:
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                    cell.font = self.normal_font
                
                row += 1
                no += 1
        
        # Summary row
        row += 1
        ws.merge_cells(f'A{row}:J{row}')
        cell = ws[f'A{row}']
        cell.value = "TOTAL VOLUME STRUKTUR"
        cell.font = Font(name='Calibri', size=11, bold=True)
        cell.fill = self.subheader_fill
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.border = self.thin_border
        
        # Total formula
        ws.cell(row=row, column=11).value = f"=SUM(K6:K{row-1})"
        ws.cell(row=row, column=11).font = Font(name='Calibri', size=11, bold=True)
        ws.cell(row=row, column=11).border = self.thin_border
        
        return ws
    
    def create_arsitektur_sheet(self):
        """Buat sheet untuk pekerjaan arsitektur"""
        ws = self.wb.create_sheet("ARSITEKTUR", 1)
        
        # Title
        ws.merge_cells('A1:J1')
        cell = ws['A1']
        cell.value = "VOLUME PEKERJAAN ARSITEKTUR - DARI GAMBAR DED"
        cell.font = Font(name='Calibri', size=14, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = self.header_fill
        
        # Header columns
        headers = ['No', 'Item Pekerjaan', 'Lokasi/Keterangan', 'Panjang (m)', 'Lebar (m)', 'Tinggi (m)', 'Jumlah', 'Satuan', 'Volume', 'Rumus/Cara Hitung']
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col_num)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.thin_border
        
        # Set column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 10
        ws.column_dimensions['I'].width = 15
        ws.column_dimensions['J'].width = 30
        
        row = 6
        categories = [
            {
                'name': 'A. PEKERJAAN DINDING',
                'items': [
                    'Pas. Dinding Bata Merah 1/2 Bata',
                    'Pas. Dinding Bata Merah 1 Bata',
                    'Pas. Dinding Hebel 10cm',
                    'Pas. Roster/Ventilasi',
                    'Plesteran 1:4',
                    'Acian'
                ]
            },
            {
                'name': 'B. PEKERJAAN PINTU & JENDELA',
                'items': [
                    'Kusen Pintu Kayu Kamper',
                    'Daun Pintu Kayu Panel',
                    'Kusen Jendela Aluminium',
                    'Daun Jendela Kaca + Aluminium',
                    'Pintu Rolling Door',
                    'Bouvenlight'
                ]
            },
            {
                'name': 'C. PEKERJAAN LANTAI',
                'items': [
                    'Pas. Keramik 40x40',
                    'Pas. Keramik 50x50',
                    'Pas. Granit 60x60',
                    'Pas. Homogenous Tile',
                    'Lantai Vynil',
                    'Plint Keramik',
                    'Plint Granit'
                ]
            },
            {
                'name': 'D. PEKERJAAN PLAFON',
                'items': [
                    'Rangka Hollow Plafon',
                    'Plafon Gypsum t=9mm',
                    'Plafon Kalsiboard',
                    'Cat Plafon',
                    'List Gypsum'
                ]
            },
            {
                'name': 'E. PEKERJAAN ATAP',
                'items': [
                    'Rangka Atap Baja Ringan',
                    'Genteng Metal',
                    'Genteng Beton',
                    'Talang Air Horizontal',
                    'Pipa Talang PVC 4"',
                    'Nok Genteng'
                ]
            },
            {
                'name': 'F. PEKERJAAN PENGECATAN',
                'items': [
                    'Cat Tembok Interior',
                    'Cat Tembok Exterior',
                    'Cat Kayu/Besi',
                    'Cat Waterproofing'
                ]
            }
        ]
        
        no = 1
        for category in categories:
            ws.merge_cells(f'A{row}:J{row}')
            cell = ws[f'A{row}']
            cell.value = category['name']
            cell.font = self.kategori_font
            cell.fill = self.kategori_fill
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = self.thin_border
            row += 1
            
            for item in category['items']:
                ws.cell(row=row, column=1).value = no
                ws.cell(row=row, column=2).value = item
                formula = f"=D{row}*E{row}*F{row}*G{row}"
                ws.cell(row=row, column=9).value = formula
                
                for col in range(1, 11):
                    cell = ws.cell(row=row, column=col)
                    cell.border = self.thin_border
                    cell.alignment = Alignment(horizontal='center' if col in [1,7,8] else 'left', vertical='center')
                    cell.font = self.normal_font
                
                row += 1
                no += 1
        
        return ws
    
    def create_mep_sheet(self):
        """Buat sheet untuk pekerjaan MEP"""
        ws = self.wb.create_sheet("MEP", 2)
        
        # Title
        ws.merge_cells('A1:J1')
        cell = ws['A1']
        cell.value = "VOLUME PEKERJAAN MEP - DARI GAMBAR DED"
        cell.font = Font(name='Calibri', size=14, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = self.header_fill
        
        # Header columns
        headers = ['No', 'Item Pekerjaan', 'Spesifikasi', 'Panjang (m)', 'Lebar (m)', 'Tinggi (m)', 'Jumlah', 'Satuan', 'Volume', 'Keterangan']
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col_num)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.thin_border
        
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 10
        ws.column_dimensions['I'].width = 15
        ws.column_dimensions['J'].width = 30
        
        row = 6
        categories = [
            {
                'name': 'A. PEKERJAAN LISTRIK',
                'items': [
                    'Panel Box',
                    'Kabel NYM 3x2.5mm',
                    'Saklar Tunggal',
                    'Stop Kontak',
                    'Lampu LED 18W',
                    'Lampu TL 36W',
                    'Conduit Pipe'
                ]
            },
            {
                'name': 'B. PEKERJAAN PLUMBING',
                'items': [
                    'Pipa Air Bersih PVC 1/2"',
                    'Pipa Air Bersih PVC 3/4"',
                    'Pipa Air Kotor PVC 4"',
                    'Pipa Air Kotor PVC 3"',
                    'Floor Drain',
                    'Kran Air',
                    'Septictank'
                ]
            },
            {
                'name': 'C. PEKERJAAN AC & MECHANICAL',
                'items': [
                    'AC Split 1 PK',
                    'AC Split 2 PK',
                    'Ducting AC',
                    'Exhaust Fan',
                    'Pompa Air'
                ]
            },
            {
                'name': 'D. SANITAIR',
                'items': [
                    'Closet Duduk',
                    'Closet Jongkok',
                    'Wastafel',
                    'Urinoir',
                    'Shower'
                ]
            }
        ]
        
        no = 1
        for category in categories:
            ws.merge_cells(f'A{row}:J{row}')
            cell = ws[f'A{row}']
            cell.value = category['name']
            cell.font = self.kategori_font
            cell.fill = self.kategori_fill
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = self.thin_border
            row += 1
            
            for item in category['items']:
                ws.cell(row=row, column=1).value = no
                ws.cell(row=row, column=2).value = item
                formula = f"=D{row}*E{row}*F{row}*G{row}"
                ws.cell(row=row, column=9).value = formula
                
                for col in range(1, 11):
                    cell = ws.cell(row=row, column=col)
                    cell.border = self.thin_border
                    cell.alignment = Alignment(horizontal='center' if col in [1,7,8] else 'left', vertical='center')
                    cell.font = self.normal_font
                
                row += 1
                no += 1
        
        return ws
    
    def create_panduan_sheet(self):
        """Buat sheet panduan pengisian"""
        ws = self.wb.create_sheet("PANDUAN", 3)
        
        ws.merge_cells('A1:F1')
        cell = ws['A1']
        cell.value = "PANDUAN PENGISIAN TEMPLATE VOLUME"
        cell.font = Font(name='Calibri', size=14, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        panduan_text = [
            "",
            "CARA MENGISI TEMPLATE:",
            "",
            "1. BUKA GAMBAR DED",
            "   - Gunakan AutoCAD, DWG TrueView, atau viewer DWG lainnya",
            "   - Perhatikan detail dimensi, layer, dan keterangan",
            "",
            "2. IDENTIFIKASI ITEM PEKERJAAN",
            "   - Lihat gambar denah, potongan, dan detail",
            "   - Catat setiap item yang terlihat",
            "   - Perhatikan as/grid untuk lokasi",
            "",
            "3. UKUR/BACA DIMENSI",
            "   - Panjang (m): Baca dari gambar",
            "   - Lebar (m): Baca dari gambar",
            "   - Tinggi/Tebal (m): Baca dari gambar atau keterangan",
            "   - Jumlah: Hitung berapa banyak item yang sama",
            "",
            "4. ISI KE TEMPLATE",
            "   - Kolom Lokasi: As/Grid atau nama ruangan",
            "   - Kolom P, L, T: Isi sesuai ukuran gambar (dalam meter)",
            "   - Kolom Jumlah: Isi jumlah item yang sama",
            "   - Kolom Satuan: m3 (kubik), m2 (persegi), m1 (meter), unit",
            "   - Kolom Volume: OTOMATIS terhitung (P x L x T x Jumlah)",
            "   - Kolom Rumus: Jelaskan cara hitungnya jika berbeda",
            "",
            "5. RUMUS VOLUME STANDAR:",
            "   • Beton (m3) = Panjang x Lebar x Tinggi x Jumlah",
            "   • Dinding (m2) = Panjang x Tinggi x Jumlah",
            "   • Lantai (m2) = Panjang x Lebar",
            "   • Pipa (m1) = Panjang total",
            "   • Pintu/Jendela (unit) = Jumlah",
            "",
            "6. CONTOH PENGISIAN:",
            "   Kolom K1 (20x30) di As A-1 s/d A-5:",
            "   - Item: Kolom Utama K-300 (20/30)",
            "   - Lokasi: As A-1 s/d A-5",
            "   - Panjang: 0.20 m",
            "   - Lebar: 0.30 m",
            "   - Tinggi: 4.00 m (tinggi lantai)",
            "   - Jumlah: 5 (jumlah kolom)",
            "   - Satuan: m3",
            "   - Volume: =0.20*0.30*4.00*5 = 1.20 m3",
            "",
            "7. TIPS MEMBACA GAMBAR:",
            "   ✓ Perhatikan skala gambar",
            "   ✓ Lihat tabel/schedule jika ada",
            "   ✓ Cek gambar detail untuk spesifikasi",
            "   ✓ Perhatikan keterangan/notes",
            "   ✓ Pastikan satuan konsisten (gunakan meter)",
            "",
            "8. SETELAH SELESAI:",
            "   ✓ Simpan file dengan nama: 'Volume_dari_Gambar.xlsx'",
            "   ✓ Jalankan script perbandingan",
            "   ✓ Script akan membandingkan dengan RAB",
            "   ✓ Hasil analisis akan otomatis dibuat",
            "",
            "CATATAN PENTING:",
            "• Isi semua item yang terlihat di gambar",
            "• Jika ragu, beri keterangan di kolom 'Rumus/Cara Hitung'",
            "• Volume akan otomatis terhitung dengan formula",
            "• Tidak perlu isi kolom volume manual (sudah ada formula)",
        ]
        
        for i, text in enumerate(panduan_text, 2):
            ws.cell(row=i, column=1).value = text
            if text.startswith(('1.', '2.', '3.', '4.', '5.', '6.', '7.', '8.')):
                ws.cell(row=i, column=1).font = Font(name='Calibri', size=11, bold=True, color="0070C0")
            elif "CARA MENGISI" in text or "CATATAN" in text:
                ws.cell(row=i, column=1).font = Font(name='Calibri', size=12, bold=True)
            else:
                ws.cell(row=i, column=1).font = Font(name='Calibri', size=10)
        
        ws.column_dimensions['A'].width = 80
        
        return ws
    
    def generate(self):
        """Generate template Excel lengkap"""
        print("\n" + "="*70)
        print("MEMBUAT TEMPLATE EXCEL UNTUK INPUT VOLUME")
        print("="*70)
        
        # Remove default sheet
        if 'Sheet' in self.wb.sheetnames:
            self.wb.remove(self.wb['Sheet'])
        
        print("\n✓ Membuat sheet STRUKTUR...")
        self.create_struktur_sheet()
        
        print("✓ Membuat sheet ARSITEKTUR...")
        self.create_arsitektur_sheet()
        
        print("✓ Membuat sheet MEP...")
        self.create_mep_sheet()
        
        print("✓ Membuat sheet PANDUAN...")
        self.create_panduan_sheet()
        
        # Save workbook
        self.wb.save(self.output_path)
        print(f"\n✓ Template berhasil dibuat: {self.output_path}")
        print("="*70)
        
        return self.output_path


if __name__ == "__main__":
    output_file = r"d:\2. NATA_PROJECTAPP\Github_RS.Sari Darma\RS-SARIDARMA\Volume_dari_Gambar_TEMPLATE.xlsx"
    
    generator = VolumeTemplateGenerator(output_file)
    generator.generate()
    
    print("\nTEMPLATE SIAP DIGUNAKAN!")
    print("\nLangkah selanjutnya:")
    print("1. Buka file gambar DED dengan viewer (AutoCAD/DWG TrueView)")
    print("2. Buka file template Excel yang baru dibuat")
    print("3. Isi data volume sesuai panduan di sheet PANDUAN")
    print("4. Simpan dengan nama: 'Volume_dari_Gambar.xlsx'")
    print("5. Jalankan script perbandingan untuk analisis\n")
