"""
Script Utama: Perbandingan Volume dari Gambar vs RAB
Menghasilkan laporan analisis lengkap dengan Excel dan visualisasi
"""

import pandas as pd
from typing import Dict, List, Tuple
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from datetime import datetime
import os


class VolumeComparator:
    """Class untuk membandingkan volume dari gambar dengan RAB"""
    
    def __init__(self, gambar_file: str, rab_files: dict):
        """
        Args:
            gambar_file: Path ke file Volume_dari_Gambar.xlsx
            rab_files: Dict dengan key 'struktur', 'arsitektur', 'mep' dan value path file RAB
        """
        self.gambar_file = gambar_file
        self.rab_files = rab_files
        self.gambar_data = {}
        self.rab_data = {}
        self.comparison_results = {}
        
        # Styles
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
        self.warning_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        self.ok_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def read_volume_gambar(self) -> Dict[str, pd.DataFrame]:
        """Baca data volume dari gambar"""
        print("\n" + "="*70)
        print("MEMBACA VOLUME DARI GAMBAR")
        print("="*70)
        
        if not os.path.exists(self.gambar_file):
            print(f"✗ File tidak ditemukan: {self.gambar_file}")
            return {}
        
        sheets = ['STRUKTUR', 'ARSITEKTUR', 'MEP']
        
        for sheet in sheets:
            try:
                df = pd.read_excel(self.gambar_file, sheet_name=sheet, header=4)
                
                # Filter data yang valid (ada volume)
                df = df[df.columns[:10]]  # Ambil 10 kolom pertama
                df.columns = ['No', 'Item', 'Lokasi', 'Panjang', 'Lebar', 'Tinggi', 'Jumlah', 'Satuan', 'Volume', 'Rumus']
                
                # Remove rows yang tidak valid
                df = df[pd.notna(df['Item'])]
                df = df[~df['Item'].astype(str).str.contains('TOTAL|PEKERJAAN|^[A-Z]\\.', na=False, regex=True)]
                df = df[pd.notna(df['Volume'])]
                df = df[df['Volume'] > 0]
                
                self.gambar_data[sheet.lower()] = df
                print(f"✓ {sheet}: {len(df)} item")
                
            except Exception as e:
                print(f"✗ Error membaca {sheet}: {e}")
                self.gambar_data[sheet.lower()] = pd.DataFrame()
        
        print("="*70)
        return self.gambar_data
    
    def read_volume_rab(self) -> Dict[str, pd.DataFrame]:
        """Baca data volume dari RAB"""
        print("\n" + "="*70)
        print("MEMBACA VOLUME DARI RAB")
        print("="*70)
        
        from rab_reader import RABReader
        
        for category, filepath in self.rab_files.items():
            if os.path.exists(filepath):
                print(f"\n→ {category.upper()}: {os.path.basename(filepath)}")
                reader = RABReader(filepath)
                data = reader.extract_data()
                
                # Convert to DataFrame
                all_items = []
                for cat, items in data.items():
                    all_items.extend(items)
                
                if all_items:
                    df = pd.DataFrame(all_items)
                    self.rab_data[category] = df
                    print(f"  ✓ {len(df)} item diekstrak")
                else:
                    self.rab_data[category] = pd.DataFrame()
                    print(f"  ✗ Tidak ada data")
            else:
                print(f"✗ {category}: File tidak ditemukan")
                self.rab_data[category] = pd.DataFrame()
        
        print("="*70)
        return self.rab_data
    
    def _is_critical_material(self, item: str) -> bool:
        """✅ Priority #8: Detect critical materials that need high matching threshold"""
        item_lower = str(item).lower()
        
        # Critical materials that must match accurately
        critical_patterns = [
            'beton k-',           # Beton with K grade (K-225, K-300, K-350, etc)
            'beton ready mix',   # Ready mix concrete
            'beton fc',           # Beton with fc (concrete strength)
            'besi diameter',      # Rebar with diameter
            'besi d',             # Besi D10, D13, D16, etc
            'besi ulir',          # Rebar (ulir)
            'besi polos',         # Plain bar
            'tulangan',           # Reinforcement
            'wiremesh',           # Wire mesh
            'kawat',              # Wire
            'semen',              # Cement
            'pasir',              # Sand
            'split',              # Gravel
            'keramik',            # Ceramic
            'granit',             # Granite
            'marmer',             # Marble
            'pipa pvc',           # PVC pipe (with diameter/schedule)
            'kabel nyyhy',        # Cable with specs
            'kabel nyy',          # Cable NYY
            'ac split',           # AC with BTU/PK
            'pompa',              # Pump with capacity
        ]
        
        for pattern in critical_patterns:
            if pattern in item_lower:
                return True
        
        return False
    
    def _get_required_threshold(self, item: str) -> float:
        """✅ Priority #8: Get required similarity threshold based on material type"""
        if self._is_critical_material(item):
            return 0.90  # 90% for critical materials (beton K-xxx, besi dia, etc)
        else:
            return 0.85  # 85% for standard materials (was 60%)
    
    def fuzzy_match_items(self, item1: str, item2: str, check_threshold: bool = True) -> float:
        """✅ Priority #8: Enhanced fuzzy matching with material-specific thresholds
        
        Args:
            item1: First item name
            item2: Second item name
            check_threshold: If True, return 0 if below threshold (for filtering)
        
        Returns:
            Similarity score (0.0 - 1.0)
        """
        from difflib import SequenceMatcher
        import re
        
        item1_clean = str(item1).lower().strip()
        item2_clean = str(item2).lower().strip()
        
        # Exact match
        if item1_clean == item2_clean:
            return 1.0
        
        # ✅ Enhanced: Check for key attribute matches in critical materials
        # For critical materials, check if key specs match (K-grade, diameter, etc)
        if self._is_critical_material(item1):
            # Extract key specs (K-225, D13, 3x2.5mm, etc)
            specs1 = set(re.findall(r'k-?\d+|d\s*\d+|fc\s*\d+|\d+x\d+\.?\d*\s*mm|\d+/\d+"', item1_clean))
            specs2 = set(re.findall(r'k-?\d+|d\s*\d+|fc\s*\d+|\d+x\d+\.?\d*\s*mm|\d+/\d+"', item2_clean))
            
            # If key specs don't match, lower similarity significantly
            if specs1 and specs2 and not specs1.intersection(specs2):
                # Different specs = different material (e.g., K-225 vs K-300, D13 vs D16)
                base_similarity = 0.5  # Force below threshold
            else:
                # Check if one contains the other
                if item1_clean in item2_clean or item2_clean in item1_clean:
                    base_similarity = 0.92  # High score for containment with matching specs
                else:
                    # Use sequence matcher
                    base_similarity = SequenceMatcher(None, item1_clean, item2_clean).ratio()
        else:
            # Standard materials: more lenient
            if item1_clean in item2_clean or item2_clean in item1_clean:
                base_similarity = 0.88  # Good score for containment
            else:
                # Use sequence matcher
                base_similarity = SequenceMatcher(None, item1_clean, item2_clean).ratio()
        
        # Apply threshold check if requested
        if check_threshold:
            required_threshold = self._get_required_threshold(item1)
            if base_similarity < required_threshold:
                return 0.0  # Below threshold = not a match
        
        return base_similarity
    
    def compare_volumes(self, category: str) -> pd.DataFrame:
        """Bandingkan volume gambar vs RAB untuk kategori tertentu"""
        print(f"\n→ Membandingkan {category.upper()}...")
        
        gambar_df = self.gambar_data.get(category, pd.DataFrame())
        rab_df = self.rab_data.get(category, pd.DataFrame())
        
        if gambar_df.empty and rab_df.empty:
            print(f"  ⚠ Tidak ada data untuk dibandingkan")
            return pd.DataFrame()
        
        comparison = []
        
        # Items dari gambar
        for _, gambar_row in gambar_df.iterrows():
            item_gambar = gambar_row['Item']
            vol_gambar = float(gambar_row['Volume']) if pd.notna(gambar_row['Volume']) else 0
            satuan_gambar = gambar_row['Satuan'] if pd.notna(gambar_row['Satuan']) else ''
            
            # Cari match di RAB (✅ Priority #8: now using enhanced matching)
            best_match = None
            best_similarity = 0.0  # Start from 0, threshold applied in fuzzy_match_items
            
            if not rab_df.empty:
                for _, rab_row in rab_df.iterrows():
                    similarity = self.fuzzy_match_items(item_gambar, rab_row['item'], check_threshold=True)
                    if similarity > best_similarity and similarity > 0:
                        best_similarity = similarity
                        best_match = rab_row
            
            if best_match is not None:
                vol_rab = float(best_match['volume']) if pd.notna(best_match['volume']) else 0
                selisih = vol_gambar - vol_rab
                selisih_persen = (selisih / vol_rab * 100) if vol_rab > 0 else 0
                
                # ✅ Priority #8: Add price validation
                harga_gambar = float(gambar_row.get('Harga Satuan', 0)) if pd.notna(gambar_row.get('Harga Satuan')) else 0
                harga_rab = float(best_match.get('harga_satuan', 0)) if pd.notna(best_match.get('harga_satuan')) else 0
                price_diff_pct = 0
                if harga_rab > 0 and harga_gambar > 0:
                    price_diff_pct = abs((harga_gambar - harga_rab) / harga_rab * 100)
                
                # Status with warnings
                status = 'MATCH'
                
                # ✅ Priority #8: Warn if similarity < 90% for critical materials
                if self._is_critical_material(item_gambar) and best_similarity < 0.90:
                    status = 'MATCH ⚠️ REVIEW (Critical material <90%)'
                
                # ✅ Priority #8: Warn if price difference > 20%
                if price_diff_pct > 20:
                    status = f'MATCH ⚠️ PRICE DIFF {price_diff_pct:.0f}%'
                
                if abs(selisih_persen) > 10:
                    status = 'SELISIH BESAR'
                elif abs(selisih_persen) > 5:
                    status = 'SELISIH KECIL'
                
                comparison.append({
                    'Item': item_gambar,
                    'Item RAB': best_match['item'],
                    'Satuan': satuan_gambar,
                    'Volume Gambar': vol_gambar,
                    'Volume RAB': vol_rab,
                    'Selisih': selisih,
                    'Selisih %': selisih_persen,
                    'Status': status,
                    'Similarity': best_similarity
                })
            else:
                comparison.append({
                    'Item': item_gambar,
                    'Item RAB': 'TIDAK DITEMUKAN',
                    'Satuan': satuan_gambar,
                    'Volume Gambar': vol_gambar,
                    'Volume RAB': 0,
                    'Selisih': vol_gambar,
                    'Selisih %': 100,
                    'Status': 'HANYA DI GAMBAR',
                    'Similarity': 0
                })
        
        # Items yang ada di RAB tapi tidak di gambar
        if not rab_df.empty:
            for _, rab_row in rab_df.iterrows():
                item_rab = rab_row['item']
                
                # Check if already matched (✅ Priority #8: use threshold checking)
                already_matched = any(
                    self.fuzzy_match_items(item_rab, comp['Item'], check_threshold=True) > 0
                    for comp in comparison
                )
                
                if not already_matched:
                    vol_rab = float(rab_row['volume']) if pd.notna(rab_row['volume']) else 0
                    
                    comparison.append({
                        'Item': 'TIDAK DITEMUKAN',
                        'Item RAB': item_rab,
                        'Satuan': rab_row.get('satuan', ''),
                        'Volume Gambar': 0,
                        'Volume RAB': vol_rab,
                        'Selisih': -vol_rab,
                        'Selisih %': -100,
                        'Status': 'HANYA DI RAB',
                        'Similarity': 0
                    })
        
        if comparison:
            result_df = pd.DataFrame(comparison)
            print(f"  ✓ {len(result_df)} item dibandingkan")
            
            # Statistics
            match_count = len(result_df[result_df['Status'] == 'MATCH'])
            selisih_count = len(result_df[result_df['Status'].str.contains('SELISIH')])
            gambar_only = len(result_df[result_df['Status'] == 'HANYA DI GAMBAR'])
            rab_only = len(result_df[result_df['Status'] == 'HANYA DI RAB'])
            
            print(f"    - Match: {match_count}")
            print(f"    - Selisih: {selisih_count}")
            print(f"    - Hanya di Gambar: {gambar_only}")
            print(f"    - Hanya di RAB: {rab_only}")
            
            return result_df
        
        return pd.DataFrame()
    
    def generate_report(self, output_file: str):
        """Generate laporan perbandingan dalam Excel"""
        print("\n" + "="*70)
        print("MEMBUAT LAPORAN PERBANDINGAN")
        print("="*70)
        
        wb = Workbook()
        wb.remove(wb.active)
        
        # Sheet ringkasan
        ws_summary = wb.create_sheet("RINGKASAN", 0)
        self.create_summary_sheet(ws_summary)
        
        # Sheet per kategori
        for category in ['struktur', 'arsitektur', 'mep']:
            if category in self.comparison_results and not self.comparison_results[category].empty:
                ws = wb.create_sheet(category.upper(), len(wb.sheetnames))
                self.create_comparison_sheet(ws, category)
        
        # Save
        wb.save(output_file)
        print(f"\n✓ Laporan berhasil dibuat: {output_file}")
        print("="*70)
    
    def create_summary_sheet(self, ws):
        """Buat sheet ringkasan"""
        # Title
        ws.merge_cells('A1:G1')
        cell = ws['A1']
        cell.value = "LAPORAN PERBANDINGAN VOLUME GAMBAR VS RAB"
        cell.font = Font(name='Calibri', size=16, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = self.header_fill
        
        # Info
        row = 3
        info_data = [
            ('Project:', 'RS Sari Dharma'),
            ('Tanggal Analisis:', datetime.now().strftime("%d-%m-%Y %H:%M")),
            ('File Gambar:', os.path.basename(self.gambar_file)),
        ]
        
        for label, value in info_data:
            ws.cell(row=row, column=1).value = label
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.cell(row=row, column=2).value = value
            row += 1
        
        # Summary table
        row += 2
        ws.merge_cells(f'A{row}:G{row}')
        cell = ws[f'A{row}']
        cell.value = "RINGKASAN PERBANDINGAN"
        cell.font = Font(name='Calibri', size=12, bold=True)
        cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        
        row += 1
        headers = ['Kategori', 'Total Item', 'Match', 'Selisih', 'Hanya Gambar', 'Hanya RAB', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.thin_border
            cell.alignment = Alignment(horizontal='center')
        
        row += 1
        for category in ['struktur', 'arsitektur', 'mep']:
            if category in self.comparison_results:
                df = self.comparison_results[category]
                if not df.empty:
                    match_count = len(df[df['Status'] == 'MATCH'])
                    selisih_count = len(df[df['Status'].str.contains('SELISIH', na=False)])
                    gambar_only = len(df[df['Status'] == 'HANYA DI GAMBAR'])
                    rab_only = len(df[df['Status'] == 'HANYA DI RAB'])
                    total = len(df)
                    
                    status = '✓ OK' if (gambar_only + rab_only) == 0 else '⚠ PERLU REVIEW'
                    
                    ws.cell(row=row, column=1).value = category.upper()
                    ws.cell(row=row, column=2).value = total
                    ws.cell(row=row, column=3).value = match_count
                    ws.cell(row=row, column=4).value = selisih_count
                    ws.cell(row=row, column=5).value = gambar_only
                    ws.cell(row=row, column=6).value = rab_only
                    ws.cell(row=row, column=7).value = status
                    
                    # Apply styling
                    for col in range(1, 8):
                        cell = ws.cell(row=row, column=col)
                        cell.border = self.thin_border
                        if gambar_only > 0 or rab_only > 0:
                            cell.fill = self.warning_fill
                        else:
                            cell.fill = self.ok_fill
                    
                    row += 1
        
        # Column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
    
    def create_comparison_sheet(self, ws, category: str):
        """Buat sheet detail perbandingan per kategori"""
        df = self.comparison_results[category]
        
        # Title
        ws.merge_cells('A1:I1')
        cell = ws['A1']
        cell.value = f"PERBANDINGAN VOLUME {category.upper()}"
        cell.font = Font(name='Calibri', size=14, bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = self.header_fill
        
        # Headers
        headers = ['No', 'Item Gambar', 'Item RAB', 'Satuan', 'Volume Gambar', 'Volume RAB', 'Selisih', 'Selisih %', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.thin_border
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
        
        # Data
        for idx, row_data in df.iterrows():
            row = idx + 4
            ws.cell(row=row, column=1).value = idx + 1
            ws.cell(row=row, column=2).value = row_data['Item']
            ws.cell(row=row, column=3).value = row_data['Item RAB']
            ws.cell(row=row, column=4).value = row_data['Satuan']
            ws.cell(row=row, column=5).value = row_data['Volume Gambar']
            ws.cell(row=row, column=6).value = row_data['Volume RAB']
            ws.cell(row=row, column=7).value = row_data['Selisih']
            ws.cell(row=row, column=8).value = row_data['Selisih %']
            ws.cell(row=row, column=8).number_format = '0.00'
            ws.cell(row=row, column=9).value = row_data['Status']
            
            # Styling
            for col in range(1, 10):
                cell = ws.cell(row=row, column=col)
                cell.border = self.thin_border
                
                # Color coding based on status
                if row_data['Status'] in ['HANYA DI GAMBAR', 'HANYA DI RAB', 'SELISIH BESAR']:
                    cell.fill = self.warning_fill
                elif row_data['Status'] == 'MATCH':
                    cell.fill = self.ok_fill
        
        # Column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 20
    
    def run_comparison(self, output_file: str):
        """Jalankan proses perbandingan lengkap"""
        print("\n" + "="*70)
        print("ANALISIS PERBANDINGAN VOLUME GAMBAR VS RAB")
        print("="*70)
        print(f"Timestamp: {datetime.now().strftime('%d-%m-%Y %H:%M:%S')}")
        print("="*70)
        
        # Read data
        self.read_volume_gambar()
        self.read_volume_rab()
        
        # Compare
        print("\n" + "="*70)
        print("PROSES PERBANDINGAN")
        print("="*70)
        
        for category in ['struktur', 'arsitektur', 'mep']:
            result = self.compare_volumes(category)
            self.comparison_results[category] = result
        
        # Generate report
        self.generate_report(output_file)
        
        print("\n" + "="*70)
        print("✓ ANALISIS SELESAI")
        print("="*70)
        print(f"\nLaporan tersimpan di: {output_file}")
        print("\nSilakan buka file Excel untuk melihat hasil detail perbandingan.")
        print("="*70 + "\n")


if __name__ == "__main__":
    # Configuration
    gambar_file = r"d:\2. NATA_PROJECTAPP\Github_RS.Sari Darma\RS-SARIDARMA\Volume_dari_Gambar.xlsx"
    
    rab_files = {
        'struktur': r"d:\2. NATA_PROJECTAPP\Github_RS.Sari Darma\RS-SARIDARMA\rab\str\BOQ-Dokumen Struktur.xlsx",
        'arsitektur': r"d:\2. NATA_PROJECTAPP\Github_RS.Sari Darma\RS-SARIDARMA\rab\ars\ANALISA VOLUME PEK  ARSITEKTUR.xlsx",
        'mep': r"d:\2. NATA_PROJECTAPP\Github_RS.Sari Darma\RS-SARIDARMA\rab\mep\RAB MEP RS SARI DARMA 17 APRIL 2025.pdf"
    }
    
    output_file = r"d:\2. NATA_PROJECTAPP\Github_RS.Sari Darma\RS-SARIDARMA\LAPORAN_PERBANDINGAN_VOLUME.xlsx"
    
    # Run comparison
    comparator = VolumeComparator(gambar_file, rab_files)
    comparator.run_comparison(output_file)
