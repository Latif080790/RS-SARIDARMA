import openpyxl

wb = openpyxl.load_workbook('output/volumes/Volume_dari_Gambar_AUTO.xlsx')
ws = wb['MEP']

print('\n' + '='*80)
print('HASIL TEXT CLEANING - MEP SHEET')
print('='*80)
print(f'{"No":<5} {"Item Description":<60} {"Volume":<10}')
print('-'*80)

for r in range(6, 30):
    no = ws.cell(r, 1).value
    item = ws.cell(r, 3).value
    volume = ws.cell(r, 11).value
    
    if item:
        print(f'{str(no or ""):<5} {str(item)[:55]:<60} {str(volume or 0):<10}')

wb.close()
print('='*80)
