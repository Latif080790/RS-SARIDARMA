"""
Detailed Verification - Check All Extracted Items
"""
import openpyxl

print("\n" + "="*100)
print("DETAILED VERIFICATION - ALL EXTRACTED ITEMS")
print("="*100)

wb = openpyxl.load_workbook('output/volumes/Volume_dari_Gambar_AUTO.xlsx')

# Check MEP Sheet
print("\n" + "="*100)
print("SHEET: MEP")
print("="*100)
ws = wb['MEP']
mep_count = 0
print(f"{'No':<5} {'Kode':<12} {'Item':<50} {'P':<8} {'L':<8} {'T':<8} {'Vol':<10}")
print("-"*100)
for r in range(5, 100):
    item = ws.cell(r, 3).value
    if item and str(item).strip():
        no = ws.cell(r, 1).value
        kode = ws.cell(r, 2).value
        p = ws.cell(r, 6).value
        l = ws.cell(r, 7).value
        t = ws.cell(r, 8).value
        vol = ws.cell(r, 11).value
        
        item_str = str(item)[:47] + "..." if len(str(item)) > 47 else str(item)
        print(f"{str(no or ''):<5} {str(kode or ''):<12} {item_str:<50} {str(p or '-'):<8} {str(l or '-'):<8} {str(t or '-'):<8} {str(vol or '-'):<10}")
        mep_count += 1

print(f"\n✓ Total MEP items: {mep_count}")

# Check STRUKTUR Sheet
print("\n" + "="*100)
print("SHEET: STRUKTUR")
print("="*100)
ws = wb['STRUKTUR']
str_count = 0
print(f"{'No':<5} {'Kode':<12} {'Item':<50} {'P':<8} {'L':<8} {'T':<8} {'Vol':<10}")
print("-"*100)
for r in range(5, 100):
    item = ws.cell(r, 3).value
    if item and str(item).strip():
        no = ws.cell(r, 1).value
        kode = ws.cell(r, 2).value
        p = ws.cell(r, 6).value
        l = ws.cell(r, 7).value
        t = ws.cell(r, 8).value
        vol = ws.cell(r, 11).value
        
        item_str = str(item)[:47] + "..." if len(str(item)) > 47 else str(item)
        print(f"{str(no or ''):<5} {str(kode or ''):<12} {item_str:<50} {str(p or '-'):<8} {str(l or '-'):<8} {str(t or '-'):<8} {str(vol or '-'):<10}")
        str_count += 1
        
        if str_count >= 10:  # Show first 10 only
            print("... (showing first 10 items)")
            break

# Count all
ws = wb['STRUKTUR']
str_total = sum(1 for r in range(5, 100) if ws.cell(r, 3).value and str(ws.cell(r, 3).value).strip())
print(f"\n✓ Total STRUKTUR items: {str_total}")

# Check ARSITEKTUR Sheet
print("\n" + "="*100)
print("SHEET: ARSITEKTUR")
print("="*100)
ws = wb['ARSITEKTUR']
ars_count = sum(1 for r in range(5, 100) if ws.cell(r, 3).value and str(ws.cell(r, 3).value).strip())
print(f"✓ Total ARSITEKTUR items: {ars_count}")

wb.close()

print("\n" + "="*100)
print("SUMMARY")
print("="*100)
print(f"MEP: {mep_count} items")
print(f"STRUKTUR: {str_total} items")
print(f"ARSITEKTUR: {ars_count} items")
print(f"TOTAL: {mep_count + str_total + ars_count} items extracted")
print("="*100)
